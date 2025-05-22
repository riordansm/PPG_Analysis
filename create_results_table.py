import os
import pickle
import pandas as pd
import re
import argparse
from pathlib import Path

def extract_pattern_info(filename):
    """Extract pattern information from filename
    
    Expected filename format: initials_testcode_date_pattern_LEDs_DRV
    Example: SR_ZTNU_051525_2A_123_DRV1_1.avi
    
    Returns:
        dict: Dictionary with extracted pattern information
    """
    pattern_info = {
        'LEDs_on': None,
        'DRV_value': None,
        'initials': None,
        'test_code': None,
        'date': None,
        'pattern': None,
        'filename': filename
    }
    
    # Remove .avi extension if present
    clean_filename = filename.replace('.avi', '')
    
    # Split by underscore
    parts = clean_filename.split('_')
    
    if len(parts) >= 6:
        # Format: initials_testcode_date_pattern_LEDs_DRV
        pattern_info['initials'] = parts[0]
        pattern_info['test_code'] = parts[1]
        pattern_info['date'] = parts[2]
        pattern_info['pattern'] = parts[3]
        
        # LEDs_on is the 5th part (index 4)
        try:
            pattern_info['LEDs_on'] = int(parts[4])
        except ValueError:
            pass
            
        # DRV_value - handle DRV1 format by extracting number after DRV
        drv_part = parts[5]
        drv_match = re.search(r'DRV(\d+)', drv_part)
        if drv_match:
            pattern_info['DRV_value'] = int(drv_match.group(1))
        else:
            # If no DRV prefix, try to parse as number
            try:
                pattern_info['DRV_value'] = int(drv_part)
            except ValueError:
                pass
    
    # Fallback: try original patterns for backward compatibility
    if pattern_info['LEDs_on'] is None or pattern_info['DRV_value'] is None:
        # Try to extract LEDs_on and DRV values from filename
        # Pattern: LEDs_on_XX_DRV_YY
        led_drv_match = re.search(r'LEDs_on_(\d+)_DRV_(\d+)', filename)
        if led_drv_match:
            pattern_info['LEDs_on'] = int(led_drv_match.group(1))
            pattern_info['DRV_value'] = int(led_drv_match.group(2))
            return pattern_info
        
        # Alternative pattern: LED_XX_DRV_YY
        led_drv_match2 = re.search(r'LED_(\d+)_DRV_(\d+)', filename)
        if led_drv_match2:
            pattern_info['LEDs_on'] = int(led_drv_match2.group(1))
            pattern_info['DRV_value'] = int(led_drv_match2.group(2))
            return pattern_info
        
        # Try to extract just numbers if pattern is different
        numbers = re.findall(r'\d+', filename)
        if len(numbers) >= 2:
            # For the fallback, use last two numbers as LEDs_on and DRV_value
            pattern_info['LEDs_on'] = int(numbers[-2])
            pattern_info['DRV_value'] = int(numbers[-1])
    
    return pattern_info

def load_results_pkl(filepath):
    """Load results from a .result.pkl file
    
    Returns:
        dict: Dictionary with analysis results or None if error
    """
    try:
        with open(filepath, 'rb') as f:
            results = pickle.load(f)
        return results
    except Exception as e:
        print(f"Error loading {filepath}: {str(e)}")
        return None

def find_results_files(root_directory):
    """Recursively find all .result.pkl files
    
    Args:
        root_directory (str): Root directory to search
        
    Returns:
        list: List of paths to .result.pkl files
    """
    results_files = []
    root_path = Path(root_directory)
    
    # Recursively search for .result.pkl files
    for pkl_file in root_path.rglob("*.result.pkl"):
        results_files.append(pkl_file)
    
    return results_files

def create_results_table(root_directory, output_csv="ppg_results_summary.csv"):
    """Create a comprehensive table of all PPG analysis results
    
    Args:
        root_directory (str): Root directory to search for results
        output_csv (str): Output CSV filename
    """
    print(f"Searching for results files in: {root_directory}")
    
    # Find all results files
    results_files = find_results_files(root_directory)
    
    if not results_files:
        print("No .result.pkl files found!")
        return
    
    print(f"Found {len(results_files)} results files")
    
    # Collect all data
    all_data = []
    
    for results_file in results_files:
        print(f"Processing: {results_file}")
        
        # Extract directory and filename info
        directory = results_file.parent.name
        full_filename = results_file.name
        
        # Remove .result.pkl to get base filename
        base_filename = full_filename.replace('.result.pkl', '')
        
        # Extract pattern information from filename
        pattern_info = extract_pattern_info(base_filename)
        
        # Load results data
        results_data = load_results_pkl(results_file)
        
        if results_data is None:
            continue
        
        # Combine all information
        row_data = {
            'Directory': directory,
            'Filename': base_filename,
            'Full_Path': str(results_file),
            'Initials': pattern_info['initials'],
            'Test_Code': pattern_info['test_code'],
            'Date': pattern_info['date'],
            'Pattern': pattern_info['pattern'],
            'LEDs_on': pattern_info['LEDs_on'],
            'DRV_value': pattern_info['DRV_value'],
            'Min_BPM': results_data.get('Min', None),
            'Max_BPM': results_data.get('Max', None),
            'Range_BPM': results_data.get('Range', None),
            'Max_Pxx': results_data.get('Max Pxx', None),
            'Mean_Pxx': results_data.get('Mean Pxx', None)
        }
        
        all_data.append(row_data)
    
    # Create DataFrame
    df = pd.DataFrame(all_data)
    
    # Add quality assessment
    df['Data_Quality'] = df['Range_BPM'].apply(
        lambda x: 'Good' if pd.notna(x) and x <= 10 else 'Poor' if pd.notna(x) else 'Unknown'
    )
    
    # Reorder columns for better readability
    column_order = [
        'Pattern', 'LEDs_on', 'DRV_value', 'Min_BPM', 'Max_BPM', 
        'Range_BPM', 'Data_Quality', 'Max_Pxx', 'Mean_Pxx'
    ]
    
    # Only include columns that exist in the dataframe
    available_columns = [col for col in column_order if col in df.columns]
    df = df[available_columns]
    
    # Sort by Pattern and DRV_value for better organization
    if 'Pattern' in df.columns and 'DRV_value' in df.columns:
        df = df.sort_values(['Pattern', 'DRV_value'], na_position='last')
    
    # Create full path for output CSV in the search directory
    root_path = Path(root_directory)
    output_path = root_path / output_csv
    
    # Save to CSV in the search directory
    df.to_csv(output_path, index=False)
    
    # Also create an Excel file with highlighting for "Good" data quality
    excel_filename = output_csv.replace('.csv', '.xlsx')
    excel_path = root_path / excel_filename
    
    try:
        # Create Excel file with conditional formatting
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='PPG Results', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['PPG Results']
            
            # Create a green fill for "Good" data quality
            from openpyxl.styles import PatternFill
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            
            # Find the Data_Quality column
            data_quality_col = None
            for col_num, col_name in enumerate(df.columns, 1):
                if col_name == 'Data_Quality':
                    data_quality_col = col_num
                    break
            
            # Apply highlighting to rows with "Good" data quality
            if data_quality_col:
                for row_num in range(2, len(df) + 2):  # Start from row 2 (after header)
                    cell_value = worksheet.cell(row=row_num, column=data_quality_col).value
                    if cell_value == 'Good':
                        # Highlight the entire row
                        for col_num in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row_num, column=col_num).fill = green_fill
        
        print(f"Excel file with highlighting saved to: {excel_path}")
        
    except ImportError:
        print("Note: openpyxl not installed. Excel file with highlighting not created.")
        print("To install: pip install openpyxl")
    except Exception as e:
        print(f"Error creating Excel file: {str(e)}")
    
    print(f"\nResults saved to: {output_path}")
    print(f"Total files processed: {len(df)}")
    
    # Print summary statistics
    print("\n=== SUMMARY STATISTICS ===")
    if 'LEDs_on' in df.columns and df['LEDs_on'].notna().any():
        print(f"LEDs_on values: {sorted(df['LEDs_on'].dropna().unique())}")
    if 'DRV_value' in df.columns and df['DRV_value'].notna().any():
        print(f"DRV values: {sorted(df['DRV_value'].dropna().unique())}")
    
    print(f"Data Quality distribution:")
    print(df['Data_Quality'].value_counts())
    
    if df['Range_BPM'].notna().any():
        print(f"\nBPM Range statistics:")
        print(f"  Mean: {df['Range_BPM'].mean():.2f}")
        print(f"  Median: {df['Range_BPM'].median():.2f}")
        print(f"  Min: {df['Range_BPM'].min():.2f}")
        print(f"  Max: {df['Range_BPM'].max():.2f}")
    
    # Show first few rows
    print(f"\n=== FIRST 10 ROWS ===")
    print(df.head(10).to_string(index=False))
    
    return df

def main():
    parser = argparse.ArgumentParser(description='Create comprehensive table of PPG analysis results')
    parser.add_argument('directory', nargs='?', default='.', 
                       help='Root directory to search for .result.pkl files (default: current directory)')
    parser.add_argument('-o', '--output', default='ppg_results_summary.csv',
                       help='Output CSV filename (default: ppg_results_summary.csv)')
    parser.add_argument('--show-patterns', action='store_true',
                       help='Show detected filename patterns for debugging')
    
    args = parser.parse_args()
    
    if args.show_patterns:
        # Debug mode: show what patterns are detected
        results_files = find_results_files(args.directory)
        print("Detected filename patterns:")
        for results_file in results_files[:10]:  # Show first 10
            base_filename = results_file.name.replace('.result.pkl', '')
            pattern_info = extract_pattern_info(base_filename)
            print(f"  {base_filename} -> LEDs_on: {pattern_info['LEDs_on']}, DRV: {pattern_info['DRV_value']}")
        return
    
    # Create the results table
    df = create_results_table(args.directory, args.output)
    
    print(f"\n✅ Complete! Results table saved to: {args.output}")

if __name__ == '__main__':
    main() 